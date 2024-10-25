import sys
import openpyxl
import requests
import csv
from google.cloud import storage
import os
from datetime import datetime, timedelta

def authentication():
    """
    potentially add keys as git.ignore for security rather than on oneDrive
    """
    service_account_path = '/mnt/c/Users/Tester/OneDrive/Christopher - Phlo/keys/phlo-sandpit-analytics-service_account.json'
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = service_account_path
    
    # code for showing contents of gcp bucket

    print("Authentication set up completed")

def get_latest_report_date():
    """
    currently scanning for monday dates when file is titled on friday date conv
    """
    current_date = datetime.now()
    days_since_monday = (current_date.weekday() - 7) % 7
    latest_monday = current_date - timedelta(days=days_since_monday)
    return latest_monday

def generate_filename(base_name, date):
    """
    comment
    """
    date_str = date.strftime("%y%m%d")
    return f"{base_name}{date_str}.xlsx"

def download_excel(url, local_filename):
    """
    comment
    """
    try:
        response = requests.get(url)
        response.raise_for_status()
        with open(local_filename, 'wb',  encoding='utf-8') as file:
            file.write(response.content)
        print(f"Downloaded {local_filename}")
        success = True
    except requests.RequestException as e:
        print(f"Error downloading file: {e}")
        success = False
    
    if success:
        return True
    else:
        return False

def modify_excel(filename):
    """
    comment
    """
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    # Add new column
    sheet['Dispenser Nominatons'] = 'Week'
    for row in range(2, sheet.max_row + 1):
        sheet[f'A{row}'] = f'Value {row}'

    # Rename pre-existing column
    old_column_name = 'Local Pharmaceutical Committee (LPC) – where blank awaiting update or DAC'
    new_column_name = 'Local Pharmaceutical Committee (LPC)'
    for cell in sheet[2]:
        if cell.value == old_column_name:
            cell.value = new_column_name
            print(f"Renamed column '{old_column_name}' to '{new_column_name}'")
            break
        
        # Add new column at the start
        sheet.insert_cols(1)
        sheet['A1'] = 'Processing Date'
        current_date = datetime.now().strftime('%Y-%m-%d')
        for row in range(2, sheet.max_row + 1):
            sheet[f'A{row}'] = current_date
        print("Added Date of NHS EPS data publication")
        
        success = True
        result = workbook
#    except Exception as e:
#        print(f"Error modifying Excel file: {e}")
#        success = False
#        result = None
#   
    if success:
        return result


def save_excel(workbook, filename):
    """
    comment
    """
    try:
        workbook.save(filename)
        print(f"Saved modified Excel file: {filename}")
        success = True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        success = False
    
    if success:
        return True
    else:
        return False

def excel_to_csv(workbook, csv_filename):
    """
    comment
    """
    try:
        sheet = workbook.active
        with open(csv_filename, 'w', encoding='utf-8', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            for row in sheet.iter_rows(values_only=True):
                csv_writer.writerow(row)
        print(f"Exported Excel to CSV: {csv_filename}")
        success = True
    except Exception as e:
        print(f"Error converting Excel to CSV: {e}")
        success = False
    
    if success:
        return True
    else:
        return False

def upload_to_gcp(bucket_name, source_file_name, destination_blob_name):
    """
    comment
    """
    try:
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(destination_blob_name)

        blob.upload_from_filename(source_file_name)
        print(f"File {source_file_name} uploaded to {destination_blob_name} in bucket {bucket_name}")
        success = True
    except Exception as e:
        print(f"Error uploading to GCP: {e}")
        success = False
    
    if success:
        return True
    else:
        return False

def main():
    """
    comment
    """
    # Set up authentication
    authentication()
    service_account_path = '/mnt/c/Users/ChristopherCampbell/Christopher - Phlo/keys/phlo-analytics-service_account.json'
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = service_account_path #debug, checking against existing files to avoid duping
    
    # Configuration
    base_url = "https://digital.nhs.uk/services/electronic-prescription-service/statistics"
    base_filename = "eps_nom_report+"
    gcp_bucket_name = "phlo-sandpit-raw-data-lake/sources/reference-data/nhs-eps-noms"

    # Get the latest report date and generate filenames
    report_date = get_latest_report_date()
    source_filename = generate_filename(base_filename, report_date)
    excel_url = f"{base_url}/{source_filename}"
    
    local_excel_filename = source_filename
    modified_excel_filename = f"modified_{source_filename}"
    local_csv_filename = f"{source_filename[:-5]}.csv"  # Replaces xlsx file type with csv
    
    gcp_csv_blob_name = f"processed_{source_filename[:-5]}.csv"

    # Download the Excel file
    if download_excel(excel_url, local_excel_filename):
        # Modify the Excel file
        modified_workbook = modify_excel(local_excel_filename)
        if modified_workbook is not None:
            # Save the modified Excel file
            if save_excel(modified_workbook, modified_excel_filename):
                # Convert Excel to CSV
                if excel_to_csv(modified_workbook, local_csv_filename):
                    # Upload CSV file to GCP
                    if upload_to_gcp(gcp_bucket_name, local_csv_filename, gcp_csv_blob_name):
                        print("Process successfully complete")
                    else:
                        print("Failed to upload CSV to GCP. Exiting Process.")
                else:
                    print("Failed to convert xlsx to CSV. Exiting Process.")
            else:
                print("Failed to save modified Excel file. Exiting Process.")
        else:
            print("Failed to modify Excel file. Exiting Process.")
    else:
        print("Failed to download Excel file. Exiting.")

    # Clean up local files
    try:
        os.remove(local_excel_filename)
        os.remove(modified_excel_filename)
        os.remove(local_csv_filename)
        print("Local files removed")
    except Exception as e:
        print(f"Error removing local files: {e}")

if __name__ == "__main__":
    main()
