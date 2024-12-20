# Primary libraries
import csv
import os
from datetime import datetime, timedelta

# 3rd party libraries
import openpyxl
import requests
from bs4 import BeautifulSoup
from google.cloud import storage

def authentication():
    """
    Set up Google Cloud auth using service account key.
    """
    try:
        service_account_path = '/mnt/c/Users/ChristopherCampbell/OneDrive - Phlo/keys/phlo-sandpit-analytics-service-account.json'
        if not os.path.exists(service_account_path):
            print(f"Service account file not found at: {service_account_path}")
            return False
            
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = service_account_path
        
        # Testing auth
        storage_client = storage.Client()
        print("Authentication successful with project:", storage_client.project)
        return True
    except Exception as e:
        print(f"Authentication error: {e}")
        return False

def list_bucket_files(bucket_name, prefix=None):
    """
    Lists all files in the specified GCP bucket.
    """
    try:
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blobs = bucket.list_blobs(prefix=prefix)
        
        files = [blob.name for blob in blobs]
        print(f"Found {len(files)} files in bucket {bucket_name} with prefix {prefix if prefix else 'none'}")
        return files
    except Exception as e:
        print(f"Error checking bucket contents: {e}")
        print(f"Bucket: {bucket_name}")
        print(f"Project: {storage_client.project}")
        return []

def check_file_exists(bucket_name, blob_name):
    """
    Check if file exists in bucket.
    """ 
    try:
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        
        exists = blob.exists()
        if exists:
            print(f"File {blob_name} already exists {bucket_name}")
        return exists
    except Exception as e:
        print(f"Error checking file existence: {e}")
        return False

def get_latest_processed_date(bucket_name, prefix=None):
    """
    Get most recent processed file date from bucket.
    Note: Looking for + in filename, not -
    """
    try:
        files = list_bucket_files(bucket_name, prefix)
        # Changed to look for the new filename format
        processed_files = [f for f in files if 'eps_nom_report+' in f]
        
        if not processed_files:
            print("No processed files found in bucket")
            return None
        
        dates = []
        for filename in processed_files:
            try:
                # Split on + instead of -
                date_str = filename.split('+')[1][:6]  # Extract YYMMDD
                date = datetime.strptime(date_str, '%y%m%d')
                dates.append(date)
            except (IndexError, ValueError):
                continue
        
        if dates:
            latest_date = max(dates)
            print(f"Latest processed file date: {latest_date.strftime('%Y-%m-%d')}")
            return latest_date
        return None
    except Exception as e:
        print(f"Error getting latest processed date: {e}")
        return None

def get_latest_report_date():
    """
    Calculate the date based on the release schedule:
    - Files are released on Mondays
    - Filenames use the previous Friday's date
    """
    current_date = datetime.now()
    current_weekday = current_date.weekday()  # Monday is 0, Friday is 4

    # Calc most recent Monday (release date)
    days_since_monday = current_weekday
    latest_monday = current_date - timedelta(days=days_since_monday)

    # Calc previous Friday (filename date)
    previous_friday = latest_monday - timedelta(days=3)  # Go back 3 days from Monday to get Friday Date

    print(f"Current date: {current_date.strftime('%Y-%m-%d')}")
    print(f"Release date (Monday): {latest_monday.strftime('%Y-%m-%d')}")
    print(f"Filename date (Previous Friday): {previous_friday.strftime('%Y-%m-%d')}")

    return previous_friday

def generate_filename(base_name, date):
    """
    Generate filename with date format.
    Note: NHS website using "-" not "+"?
    Format should be "eps_noms_report+{FridayDate}"
    """
    date_str = date.strftime("%y%m%d")
    return f"{base_name.replace('+', '-')}{date_str}.xlsx"

def download_excel(url, local_filename):
    """
    Download Excel file from NHS website.
    """
    try:
        print(f"Accessing webpage: {url}")
        response = requests.get(url)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Adjust target filename to match NHS website format
        target_file = local_filename.replace('+', '-')
        print(f"Looking for file: {target_file}")
        
        download_url = None
        for link in soup.find_all('a', href=True):
            href = link['href']
            if target_file.replace('.xlsx', '') in href:
                download_url = href if href.startswith('http') else f"https://digital.nhs.uk{href}"
                print(f"\nFound download URL: {download_url}")
                break
        
        if not download_url:
            print(f"\nCould not find download link for {target_file}")
            return False
            
        print(f"\nDownloading file from: {download_url}")
        file_response = requests.get(download_url)
        file_response.raise_for_status()
        
        with open(local_filename, 'wb') as file:
            file.write(file_response.content)
        print(f"Successfully downloaded {local_filename}")
        return True
        
    except requests.RequestException as e:
        print(f"Error downloading file: {e}")
        print(f"Attempted URL: {url}")
        return False

def modify_excel(filename):
    """
    Modify Excel file with required changes:
    - Select 'Dispenser Nominations' sheet
    - Add 'Week' column with appropriate date
    - Rename LPC column title (I1)
    - Save only this sheet
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        
        # Print all sheet names for debugging
        print("\nAll available sheets in workbook:")
        for sheet_name in workbook.sheetnames:
            print(f"- {sheet_name}")
        
        # Get the Dispenser Nominations sheet
        if 'Dispenser Nominations' not in workbook.sheetnames:
            raise Exception("Could not find 'Dispenser Nominations' sheet")
        
        # Select the correct sheet
        sheet = workbook['Dispenser Nominations']
        print(f"\nWorking with sheet: {sheet.title}")
        
        # Find last populated row in column B
        last_populated_row = 1  # Start at 1 to account for header
        for row in range(2, sheet.max_row + 1):
            if sheet[f'B{row}'].value is not None:
                last_populated_row = row
            else:
                break  # Exit loop when we find first empty cell
        
        print(f"Found last populated row in column B: {last_populated_row}")
        
        # Extract date from filename
        date_str = filename.split('-')[-1].replace('.xlsx', '')  # Get YYMMDD part
        file_date = datetime.strptime(date_str, '%y%m%d')
        formatted_date = file_date.strftime('%Y-%m-%d')
        print(f"Using date: {formatted_date}")
        
        # Insert new column A
        sheet.insert_cols(1)
        sheet['A1'] = 'Week'
        print("Added 'Week' column")
        
        # Fill date down column A to match populated rows
        print(f"Filling dates down to row {last_populated_row}")
        for row in range(2, last_populated_row + 1):
            sheet[f'A{row}'] = formatted_date
        
        print(f"Added date {formatted_date} to rows 2 through {last_populated_row}")

        # Rename LPC column (I1) after inserting new column
        old_title = sheet['I1'].value
        sheet['I1'] = 'Local Pharmaceutical Committee (LPC)'
        print(f"Renamed column I1 from '{old_title}' to 'Local Pharmaceutical Committee (LPC)'")
        
        return workbook
    except Exception as e:
        print(f"Error modifying Excel file: {e}")
        print(f"Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        return None

def save_excel(workbook, filename):
    """
    Save modified Excel workbook.
    """
    try:
        workbook.save(filename)
        print(f"Saved modified Excel file: {filename}")
        return True
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False

def excel_to_csv(workbook, csv_filename):
    """
    Convert specific Excel sheet to CSV.
    """
    try:
        sheet = workbook['Dispenser Nominations']  # Specifically use the Dispenser Nominations sheet
        with open(csv_filename, 'w', encoding='utf-8', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            for row in sheet.iter_rows(values_only=True):
                csv_writer.writerow(row)
        print(f"Exported 'Dispenser Nominations' sheet to CSV: {csv_filename}")
        return True
    except Exception as e:
        print(f"Error converting Excel to CSV: {e}")
        print(f"Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        return False

def upload_to_gcp(bucket_name, source_file_name, destination_blob_name):
    """
    Upload file to GCP bucket.
    """
    try:
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(destination_blob_name)

        blob.upload_from_filename(source_file_name)
        print(f"File {source_file_name} uploaded to {destination_blob_name} in bucket {bucket_name}")
        return True
    except Exception as e:
        print(f"Error uploading to GCP: {e}")
        return False
    
def setup_working_directory():
    """
    Create and use a specific directory for working files.
    """
    # Create directory in user's home directory
    work_dir = os.path.expanduser("~/eps_nominations_processing")
    if not os.path.exists(work_dir):
        os.makedirs(work_dir)
    
    # Change to this directory
    os.chdir(work_dir)
    print(f"\nWorking directory: {os.getcwd()}")
    return work_dir

    
def cleanup_files(files_to_remove):
    """
    Safely remove local files if they exist.
    """
    for file in files_to_remove:
        if os.path.exists(file):
            try:
                os.remove(file)
                print(f"Removed {file}")
            except Exception as e:
                print(f"Error removing {file}: {e}")
        else:
            print(f"File not found, skipping: {file}")

def main():
    """
    Main function with GCP bucket checking.
    """
    # Set up working directory first
    # work_dir = setup_working_directory()
    # print(f"Files will be processed in: {work_dir}")
    
    # Set up working directory first
    work_dir = setup_working_directory()
    print(f"Files will be processed in: {work_dir}")

    # Set up auth
    if not authentication():
        print("Authentication failed. Exiting.")
        return 
   
    # Config
    base_url = "https://digital.nhs.uk/services/electronic-prescription-service/statistics"
    base_filename = "eps_nom_report+"  # Note the + here
    gcp_bucket_name = "phlo-sandpit-raw-data-lake"
    blob_prefix = "sources/reference-data/nhs-eps-noms/"

    # Check latest file in GCP
    latest_processed_date = get_latest_processed_date(gcp_bucket_name, blob_prefix)
    
    # Get the latest report date
    report_date = get_latest_report_date()

    # Print dates for debugging
    print(f"Latest processed date: {latest_processed_date}")
    print(f"Current report date: {report_date}")
    
    # Skip if file exists
    if latest_processed_date and report_date <= latest_processed_date:
        print("Already have the latest file processed. Skipping download.")
        return
    
    # Generate filenames - Note we use - for download but + for GCP
    source_filename = generate_filename(base_filename, report_date)  # Will have - for download
    gcp_filename = source_filename.replace('-', '+')  # Convert to + for GCP storage
    
    excel_url = base_url
    
    local_excel_filename = source_filename
    modified_excel_filename = f"modified_{source_filename}"
    local_csv_filename = gcp_filename.replace('.xlsx', '.csv')  # Use + version for CSV
     # Remove 'processed_' prefix from GCP blob name
    gcp_csv_blob_name = f"{blob_prefix}{gcp_filename[:-5]}.csv"  # Use + version for GCP

    # Check if file exists
    if check_file_exists(gcp_bucket_name, gcp_csv_blob_name):
        print(f"File {gcp_csv_blob_name} already exists in GCP. Skipping processing.")
        return

    # Download and process the file
    if download_excel(excel_url, local_excel_filename):
        modified_workbook = modify_excel(local_excel_filename)
        if modified_workbook is not None:
            if save_excel(modified_workbook, modified_excel_filename):
                if excel_to_csv(modified_workbook, local_csv_filename):
                    if upload_to_gcp(gcp_bucket_name, local_csv_filename, gcp_csv_blob_name):
                        print("Process successfully complete")
                    else:
                        print("Failed to upload CSV to GCP. Exiting Process.")
                else:
                    print("Failed to convert xlsx to CSV. Exiting Process.")
            else:
                print("Failed to save modified Excel file. Exiting Process.")
        else:
            print("Failed to modify Excel file. Exiting Process.")
    else:
        print("Failed to download Excel file. Exiting.")

    """
    # Clean up
    cleanup_files([
        local_excel_filename,
        modified_excel_filename,
        local_csv_filename
    ])
    """
if __name__ == "__main__":
    main()
