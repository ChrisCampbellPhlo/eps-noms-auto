import os
import requests
import openpyxl
from google.cloud import storage

def download_excel(url, local_filename):
    response = requests.get(url)
    response.raise_for_status()
    with open(local_filename, 'wb') as file:
        file.write(response.content)
    print(f"Downloaded {local_filename}")

def modify_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    # Add new column
    sheet['A1'] = 'New Column'
    for row in range(2, sheet.max_row + 1):
        sheet[f'D{row}'] = f'Value {row}'

    # Rename pre-existing column
    old_column_name = 'Local Pharmaceutical Committee (LPC) – where blank awaiting update or DAC'
    new_column_name = 'Local Pharmaceutical Committee (LPC)'
    for cell in sheet[2]:
        if cell.value == old_column_name:
            cell.value = new_column_name
            print(f"Renamed column '{old_column_name}' to '{new_column_name}'")
            break
    
  #  workbook.save(filename)
  #  print(f"Modified {filename}")


#convert file to from xlsx to csv
def save_excel(workbook, filename):
    workbook.save(filename)
    print(f"Saved modified Excel file: {filename}")

def excel_to_csv(workbook, csv_filename):
    sheet = workbook.active
    with open(csv_filename, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        for row in sheet.iter_rows(values_only=True):
            csv_writer.writerow(row)
    print(f"Exported Excel to CSV: {csv_filename}")


# naming convention for files 
def get_latest_report_date():
    # Get the current date
    current_date = datetime.now()
    # Find most recent Monday (Report is generated on Mondays (Assumed manual upload, add in funct for delays)
    days_since_monday = (current_date.weekday() - 0) % 7
    latest_monday = current_date - timedelta(days=days_since_monday)
    return latest_monday

def generate_filename(base_name, date):
    # Format the date as required: YYMMDD
    date_str = date.strftime("%y%m%d")
    return f"{base_name}{date_str}.xlsx"

def download_excel(url, local_filename):
    response = requests.get(url)
    response.raise_for_status()
    with open(local_filename, 'wb') as file:
        file.write(response.content)
    print(f"Downloaded {local_filename}")

def modify_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active


# add authentication protocol
#play funct for checking gcp bucket
# 
def authentication():
    service_account_path = '/mnt/c/Users/ChristopherCampbell/Christopher - Phlo/keys/phlo-analytics-service_account.json'
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = service_account_path #debug, checking against existing files to avoid duping
    return

def upload_to_gcp(bucket_name, source_file_name, destination_blob_name):
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)

    blob.upload_from_filename(source_file_name)
    print(f"File {source_file_name} uploaded to {destination_blob_name} in bucket {bucket_name}")

def main():
    authentication()
    sys.exit()

    # Configuration
    excel_url = "https://digital.nhs.uk/services/electronic-prescription-service/statistics"
    local_filename = "eps_nom_report.xlsx"
    base_filename = "eps_nom_report+"
    gcp_bucket_name = "phlo-sandpit-raw-data-lake/sources/reference-data/nhs-eps-noms"
    gcp_destination_blob_name = ".xlsx"

    # Download the Excel file
    download_excel(excel_url, local_filename)

    # Modify the Excel file
    modify_excel(local_filename)

    # Upload the modified file to GCP
    upload_to_gcp(gcp_bucket_name, local_filename, gcp_destination_blob_name)

    # Clean up the local file
    os.remove(local_filename)
    print("Local file removed")

if __name__ == "__main__":
    main()